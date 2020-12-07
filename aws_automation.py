
##############################################################################################################
#                                                                                                            #
# 1. Import Packages                                                                                         #
# 2. CREDENTIALS AND REGION DECLARATION                                                                      #
# 3. VPC                                                                                                     #
# 4. SUBNET                                                                                                  #
# 5. ROUTE TABLE                                                                                             #
# 6. INTERNET GATEWAY                                                                                        #
# 7. DHCP OPTIONS SET                                                                                        #
# 8. PEERING CONNECTION                                                                                      #
# 9. CUSTOMER GATEWAY                                                                                        #
# 10.VIRTUAL PRIVATE GATEWAY                                                                                 #   
# 11.SITE-TO-SITE VPN CONNECTION                                                                             #
# 12.VPC ENDPOINTS                                                                                           #
# 13.INVENTORY                                                                                               #
# 14.DECLARING IN ARGPARSER                                                                                  #
##############################################################################################################


# 1. Import Packages
import requests
import requests_negotiate_sspi
import re
import html
import urllib
import os
import boto3
import argparse
import webbrowser
from prettytable import PrettyTable
from botocore.exceptions import ClientError
from openpyxl import load_workbook
from openpyxl.styles import Font
from pprint import pprint

# 2. Credentials and Region declaration

def connection(region):
    client = boto3.client('ec2', aws_access_key_id='AKIAVVQ5DYUON5SMOLFD',
                      aws_secret_access_key='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',
                      region_name=region)
    return client


def connection1(region):
    clients = boto3.client('sts', aws_access_key_id='AKIAVVQ5DYUON5SMOLFD',
                       aws_secret_access_key='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',
                       region_name=region)
    print("\nConnected..")
    return clients




# 3. VPC
##############################################################################################################
#       3.1 Create VPC                                                                                       #
#       3.2 Describe VPC                                                                                     #
#       3.3 Delete VPC                                                                                       #
##############################################################################################################

# 3.1 Create VPC
def vpc():
    try:
        cidr_block = input("\nIPV4 CIDR block : ")
        vpc = client.create_vpc(
            CidrBlock=cidr_block
        )
        connectionId = vpc['Vpc']['VpcId']
        print(connectionId)
        name = input("VPC Name : ")
        client.create_tags(
            Resources=[connectionId],
            Tags=[{"Key": "Name", "Value": name}]
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 3.2 Describe VPC
def describe_vpc():
    try:
        response = client.describe_vpcs()
        vpc = response['Vpcs']
        print("\nThe available Vpcs in your region\n")
        for vpcs in vpc:
            fetch = vpcs['VpcId']
            cidr = vpcs['CidrBlock']
            tags = vpcs['Tags'][0]
            print("\t\tVPC-ID : {} | CidrBlock : {} | VPC-NAME : {}".format(fetch, cidr, tags['Value']))
    except ClientError as e:
        print("\n")
        print(e)


# 3.3 Add secondary cidr
def secondary_cidr():
    try:
        describe_vpc()
        cidr_block = input("\nIPV4 CIDR block : ")
        id = input("\nVpcId : ")
        response = client.associate_vpc_cidr_block(
            CidrBlock=cidr_block,
            VpcId=id
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 3.4 Delete VPC
def delete_vpc():
    try:
        describe_vpc()
        id = input("\nVpcId : ")
        response = client.delete_vpc(
            VpcId=id,
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 4. SUBNET
##############################################################################################################
#       4.1 CREATE SUBNET                                                                                    #
#       4.2 DESCRIBE SUBNET                                                                                  #
#       4.3 DELETE SUBNET                                                                                    #
##############################################################################################################

# 4.1 Create SUBNET
def subnet():
    try:
        describe_vpc()
        zone = input("\nAvailability Zone : ")
        cidr = input("\nIPV4 CIDR Block : ")
        vpc = input("\nVPC ID : ")
        response = client.create_subnet(
            AvailabilityZone=zone,
            CidrBlock=cidr,
            VpcId=vpc,
        )
        connectionId = response['Subnet']['SubnetId']
        print("\n" + connectionId)
        name = input("SUBNET Name : ")
        client.create_tags(
            Resources=[connectionId],
            Tags=[{"Key": "Name", "Value": name}]
        )
        print("\nSuccess")
    except ClientError as e:
        print("\n")
        print(e)


# 4.2 Describe SUBNET
def describe_subnet():
    try:
        response = client.describe_subnets()
        subnet = response['Subnets']
        print("\nThe available Subnets in your region\n")
        for sub in subnet:
            fetch = sub['SubnetId']
            cidr = sub['CidrBlock']
            vpc = sub['VpcId']
            tags = sub['Tags'][0]
            print("\t\tSUBNET-ID : {} |CIDR BLOCK : {} | VPC-ID : {} | SUBNET-NAME : {}".
                  format(fetch, cidr, vpc, tags['Value']))
    except ClientError as e:
        print("\n")
        print(e)


# 4.3 Delete SUBNET
def delete_subnet():
    try:
        describe_subnet()
        id = input("\nSubnetId : ")
        response = client.delete_subnet(
            SubnetId=id,
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 5. ROUTE TABLE
##############################################################################################################
#       5.1 MODIFY NAME FOR  ROUTE TABLE                                                                     #
#       5.2 CREATE ROUTE TABLE                                                                               #
#       5.3 DESCRIBE ROUTE TABLE                                                                             #
#       5.4 CREATE ROUTE FOR INTERNET GATEWAY                                                                #
#       5.5 CREATE ROUTE FOR PEERING CONNECTION                                                              #
#       5.6 CREATE ROUTE FOR VIRTUAL PRIVATE GATEWAY                                                         #
#       5.7 SUBNET ASSOCIATION IN ROUTE TABLE                                                                #
#       5.8 DESCRIBE ROUTE                                                                                   #
#       5.9 DELETE ROUTE                                                                                     #
#       5.10 DELETE ROUTE TABLE                                                                              #
##############################################################################################################

# 5.1 MODIFY NAME FOR  ROUTE TABLE
def modify_route_table():
    try:
        describe_route_table()
        connectionId = input("\nRoute Table ID : ")
        name = input("\nRoute Table name : ")
        client.create_tags(
            Resources=[connectionId],
            Tags=[{"Key": "Name", "Value": name}]
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 5.2 CREATE ROUTE TABLE
def route_table():
    try:
        describe_vpc()
        id = input("\nEnter the VPC ID : ")
        response = client.create_route_table(
            VpcId=id
        )
        connectionId = response['RouteTable']['RouteTableId']
        print('\n' + connectionId)
        name = input("\nRoute Table name : ")
        client.create_tags(
            Resources=[connectionId],
            Tags=[{"Key": "Name", "Value": name}]
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 5.3 DESCRIBE ROUTE TABLE
def describe_route_table():
    try:
        response = client.describe_route_tables()
        #pprint(response)
        rt = response['RouteTables']
        print("\nThe available RouteTables in your region\n")
        for rtl in rt:
            fetch = rtl['RouteTableId']
            tags = rtl['Tags'][0]
            vpc = rtl['VpcId']
            print("\t\tROUTE TABLE-ID : {} | ROUTE TABLE-NAME : {} | Attached VPC : {}".
                  format(fetch, tags['Value'], vpc))
    except ClientError as e:
        print("\n")
        print(e)


# 5.4 CREATE ROUTE FOR INTERNET GATEWAY
def route_igw():
    try:
        describe_igw()
        describe_route_table()
        cd = input("\nDestination CidrBlock : ")
        gw = input("\nInternet Gateway Id : ")
        rt = input("\nRoute Table Id : ")
        response = client.create_route(
            DestinationCidrBlock=cd,
            GatewayId=gw,
            RouteTableId=rt,
        )
        print("\nRoute been created\n")
    except ClientError as e:
        print("\n")
        print(e)


# 5.5 CREATE ROUTE FOR PEERING CONNECTION
def route_vpc_peering():
    try:
        describe_vpc_peering()
        describe_route_table()
        cidr = input("\nDestinationCidrBlock : ")
        vpcp_id = input("VpcPeeringConnectionId : ")
        rt_id = input("RouteTableId : ")
        response = client.create_route(
            DestinationCidrBlock=cidr,
            VpcPeeringConnectionId=vpcp_id,
            RouteTableId=rt_id
        )
        print("\nSuccess\n")
    except ClientError as e:
        print('\n')
        print(e)


# 5.6 CREATE ROUTE FOR VIRTUAL PRIVATE GATEWAY
def route_vpn_gateway():                                                                    
    try:
        describe_vpn_gateway()
        describe_route_table()
        cidr = input("\nDestinationCidrBlock : ")
        vpg_id = input("GatewayId : ")
        rt_id = input("RouteTableId : ")
        response = client.create_route(
            DestinationCidrBlock=cidr,
            GatewayId=vpg_id,
            RouteTableId=rt_id
        )
        print("\nSuccess")
    except ClientError as e:
        print('\n')
        print(e)


# 5.7 SUBNET ASSOCIATION IN ROUTE TABLE
def attach_subnet_route_table():
    try:
        describe_route_table()
        describe_subnet()
        rt_id = input("\nRouteTableId : ")
        sn_id = input("SubnetId : ")
        response = client.associate_route_table(
            RouteTableId=rt_id,
            SubnetId=sn_id,
        )
        print("\nSuccess")
    except ClientError as e:
        print('\n')
        print(e)

# 5.8 DESCRIBE SUBNET ASSOCIATION IN ROUTE TABLE
def describe_subnet_association():
    try:
        response = client.describe_route_tables()
        route_tables = response['RouteTables']
        for association in route_tables:
            rt_id = association['RouteTableId']
            rt_name = association['Tags'][0]

            for route in association['Associations']:

                if 'SubnetId' in route:
                    associated_id = route['RouteTableAssociationId']
                    subnet_id = route['SubnetId']
                    print("\nRouteTable ID : {} | RouteTable Name : {} | AssociationId : {} | Subnet ID : {} ".format(rt_id, rt_name['Value'], associated_id, subnet_id))
    except ClientError as e:
        print('\n')
        print(e)


# 5.9 DETACH SUBNET ASSOCIATION IN ROUTE TABLE
def detach_subnet_route_table():
    try:
        describe_subnet_association()
        ass_id = input("\nAssociationId : ")
        response = client.disassociate_route_table(
            AssociationId=ass_id
        )
        print("\nSuccess\n")
    except ClientError as e:
        print('\n')
        print(e)


# 5.10 DESCRIBE ROUTE
def describe_route():
    try:
        response = client.describe_route_tables()
        route_tables = response['RouteTables']
        route_tables_count = 0;
        for association in route_tables:
            rt_id = association['RouteTableId']
            rt_name = association['Tags'][0]
            print("\nRouteTable ID : {} | RouteTable Name : {}".format(rt_id, rt_name['Value'])) 
            #print("\nRoute Table ID : "+association['RouteTableId'])
            for route in association['Routes']:
                if 'EgressOnlyInternetGatewayId' in route:
                    print("Destination Cidr Block : {} | Target : {}".format(route['DestinationCidrBlock'],route['EgressOnlyInternetGatewayId']))
                if 'GatewayId' in route:
                    print("Destination Cidr Block : {} | Target : {}".format(route['DestinationCidrBlock'],route['GatewayId']))
                if 'InstanceId' in route:
                    print("Destination Cidr Block : {} | Target : {}".format(route['DestinationCidrBlock'],route['InstanceId']))
                if 'InstanceOwnerId' in route:
                    print("Destination Cidr Block : {} | Target : {}".format(route['DestinationCidrBlock'],route['InstanceOwnerId']))
                if 'NatGatewayId' in route:
                    print("Destination Cidr Block : {} | Target : {}".format(route['DestinationCidrBlock'],route['NatGatewayId']))
                if 'TransitGatewayId' in route:
                    print("Destination Cidr Block : {} | Target : {}".format(route['DestinationCidrBlock'],route['TransitGatewayId']))
                if 'LocalGatewayId' in route:
                    print("Destination Cidr Block : {} | Target : {}".format(route['DestinationCidrBlock'],route['LocalGatewayId']))
                if 'NetworkInterfaceId' in route:
                    print("Destination Cidr Block : {} | Target : {}".format(route['DestinationCidrBlock'],route['NetworkInterfaceId']))
                if 'VpcPeeringConnectionId' in route:
                    print("Destination Cidr Block : {} | Target : {}".format(route['DestinationCidrBlock'],route['VpcPeeringConnectionId']))
                   
    except ClientError as e:
        print("\n")
        print(e)
        

# 5.11 DELETE ROUTE
def delete_route():
    try:
        describe_route()
        describe_route_table()
        cidr = input("\nDestinationCidrBlock : ")
        id = input("\nRouteTableId : ")
        response = client.delete_route(
            DestinationCidrBlock=cidr,
            RouteTableId=id
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 5.12 DELETE ROUTE TABLE
def delete_route_table():
    try:
        describe_route_table()
        id = input("\nRouteTableId : ")
        response = client.delete_route_table(
            RouteTableId=id
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 6.INTERNET GATEWAY
##############################################################################################################
#       6.1 CREATE INTERNET GATEWAY                                                                          #
#       6.2 Describe INTERNET GATEWAY                                                                        #
#       6.3 ATTACH INTERNET GATEWAY TO VPC                                                                   #
#       6.4 DETACH INTERNET GATEWAY FROM VPC                                                                 #
#       6.5 Delete INTERNET GATEWAY                                                                          #
##############################################################################################################

# 6.1 CREATE INTERNET GATEWAY
def internet_gateway():
    try:
        response = client.create_internet_gateway()
        connectionId = response['InternetGateway']['InternetGatewayId']
        print('\n' + connectionId)
        name = input("\nInternet Gateway Name : ")
        client.create_tags(
            Resources=[connectionId],
            Tags=[{"Key": "Name", "Value": name}]
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 6.2 Describe INTERNET GATEWAY
def describe_igw():
    try:
        response = client.describe_internet_gateways()
        igw = response['InternetGateways']
        print("\nThe available Internet Gateway in your region\n ")
        for igws in igw:
            fetch = igws['InternetGatewayId']
            tags = igws['Tags'][0]
            print("\t\tIGW-ID : {} | IGW-NAME : {}".format(fetch, tags['Value']))
    except ClientError as e:
        print("\n")
        print(e)


# 6.3 ATTACH INTERNET GATEWAY TO VPC
def attach_igw():
    try:
        describe_igw()
        describe_vpc()
        id = input("\nInternetGatewayId : ")
        v_id = input("\nVpcId : ")
        response = client.attach_internet_gateway(
            InternetGatewayId=id,
            VpcId=v_id
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 6.4 DETACH INTERNET GATEWAY FROM VPC
def detach_igw():
    try:
        describe_igw()
        describe_vpc()
        id = input("\nInternetGatewayId : ")
        v_id = input("\nVpcId : ")
        response = client.detach_internet_gateway(
            InternetGatewayId=id,
            VpcId=v_id
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 6.5 Delete INTERNET GATEWAY
def delete_igw():
    try:
        detach_igw()
        describe_igw()
        id = input("\nInternetGatewayId : ")
        response = client.delete_internet_gateway(
            InternetGatewayId=id
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 8.PEERING CONNECTIONS
##############################################################################################################
#       8.1 CREATE PEERING CONNECTIONS                                                                       #
#       8.2 CREATE PEERING CONNECTIONS TO ANOTHER REGION                                                     #
#       8.3 Describe PEERING CONNECTIONS                                                                     #
#       8.4 ACCEPT PEERING CONNECTIONS                                                                       #
#       8.5 REJECT PEERING CONNECTIONS                                                                       #
#       8.6 OWNER ID                                                                                         #
#       8.7 DELETE PEERING CONNECTIONS                                                                       #
##############################################################################################################

# 8.1 CREATE PEERING CONNECTIONS
def vpc_peering():
    try:
        owner_id()
        describe_vpc()
        own_id = input("\nEnter the Owner Id : ")
        vpcid = input("Enter the Requester VPC ID : ")
        vpcids = input("Enter the Accepter ID : ")
        peering = client.create_vpc_peering_connection(
            PeerOwnerId=own_id,
            PeerVpcId=vpcid,
            VpcId=vpcids,
            PeerRegion=region
        )
        connectionId = peering['VpcPeeringConnection']['VpcPeeringConnectionId']
        print(connectionId)
        name = input("Enter VPC Peering Name : ")
        client.create_tags(
            Resources=[connectionId],
            Tags=[{"Key": "Name", "Value": name}]
        )
        print("\nSuccess\n")
    except ClientError as e:
        print('\n')
        print(e)


# 8.2 CREATE PEERING CONNECTIONS TO ANOTHER REGION 
def vpc_peering_another_region():
    try:
        owner_id()
        describe_vpc()
        own_id = input("\nEnter the Owner Id : ")
        vpcid = input("Enter the Requester VPC ID : ")
        reg = input("Enter the Accepter VPC Region : ")
        vpcids = input("Enter the Accepter ID : ")
        peering = client.create_vpc_peering_connection(
            PeerOwnerId=own_id,
            PeerVpcId=vpcid,
            VpcId=vpcids,
            PeerRegion=reg
        )
        connectionId = peering['VpcPeeringConnection']['VpcPeeringConnectionId']
        print(connectionId)
        name = input("Enter VPC Peering Name : ")
        client.create_tags(
            Resources=[connectionId],
            Tags=[{"Key": "Name", "Value": name}]
        )
        print("\nSuccess\n")
    except ClientError as e:
        print('\n')
        print(e)

        
# 8.3 Describe PEERING CONNECTIONS
def describe_vpc_peering():
    try:
        response = client.describe_vpc_peering_connections()
        vpc = response['VpcPeeringConnections']
        print("\nThe available VPC Peering Connections in your region\n")
        for vpcp in vpc:
            fetch = vpcp['VpcPeeringConnectionId']
            tags = vpcp['Tags'][0]
            status = vpcp['Status']
            print("\t\tVPC Peering Connection-ID : {} | Name : {} | Status : {}".
                  format(fetch, tags['Value'], status['Code']))
    except ClientError as e:
        print('\n')
        print(e)


# 8.4 ACCEPT PEERING CONNECTIONS
def accept_peering():
    try:
        describe_vpc_peering()
        id = input("\nEnter the VPC Peering Connection Id : ")
        connection = client.accept_vpc_peering_connection(
            VpcPeeringConnectionId=id
        )
        print("\nSuccess\n")
    except ClientError as e:
        print('\n')
        print(e)


# 8.5 REJECT PEERING CONNECTIONS
def reject_peering():
    try:
        describe_vpc_peering()
        id = input("\nEnter the VPC Peering Connection Id : ")
        response = client.reject_vpc_peering_connection(
            VpcPeeringConnectionId=id
        )
        print("\nSuccess\n")
    except ClientError as e:
        print('\n')
        print(e)


# 8.6 OWNER ID
def owner_id():                                                                                      
    try:
        response = clients.get_caller_identity()
        print('\nOWNER ID : ', response['Account'])
    except ClientError as e:
        print('\n')
        print(e)


# 8.7 DELETE PEERING CONNECTIONS
def delete_vpc_peering():
    try:
        describe_vpc_peering()
        id = input("\nVpcPeeringConnectionId : ")
        response = client.delete_vpc_peering_connection(
            VpcPeeringConnectionId=id
        )
        print("\nSuccess\n")
    except ClientError as e:
        print('\n')
        print(e)


# 9.CUSTOMER GATEWAY
##############################################################################################################
#       9.1 CREATE CUSTOMER GATEWAY                                                                          #
#       9.2 Describe CUSTOMER GATEWAY                                                                        #
#       9.3 DELETE CUSTOMER GATEWAY                                                                          #
##############################################################################################################

# 9.1 CREATE CUSTOMER GATEWAY
def customer_gateway():
    try:
        ip = input("\nPublicIp : ")
        name = input("DeviceName : ")
        response = client.create_customer_gateway(
            BgpAsn=64512,
            PublicIp=ip,
            Type='ipsec.1',
            DeviceName=name,
        )
        connectionId = response['CustomerGateway']['CustomerGatewayId']
        print(connectionId)
        name = input("Customer Gateway Name : ")
        client.create_tags(
            Resources=[connectionId],
            Tags=[{"Key": "Name", "Value": name}]
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 9.2 Describe CUSTOMER GATEWAY
def describe_customer_gateway():
    try:
        response = client.describe_customer_gateways()
        cg = response['CustomerGateways']
        print("\nThe available Customer Gateway in your region\n")
        for cgs in cg:
            fetch = cgs['CustomerGatewayId']
            tags = cgs['Tags'][0]
            name = cgs['State']
            print("\t\tCustomer Gateway-ID : {} | State : {} | Customer Gateway-Name : {}".
                  format(fetch, name, tags['Value']))
    except ClientError as e:
        print("\n")
        print(e)


# 9.3 DELETE CUSTOMER GATEWAY
def delete_customer_gateway():
    try:
        describe_customer_gateway()
        id = input("\nCustomerGatewayId : ")
        response = client.delete_customer_gateway(
            CustomerGatewayId=id,
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 10.VIRTUAL PRIVATE GATEWAY
##############################################################################################################
#       10.1 CREATE VIRTUAL PRIVATE GATEWAY                                                                  #
#       10.2 Describe VIRTUAL PRIVATE GATEWAY                                                                #
#       10.3 ATTACH VIRTUAL PRIVATE GATEWAY                                                                  #
#       10.4 DETACH VIRTUAL PRIVATE GATEWAY                                                                  #
#       10.5 DELETE VIRTUAL PRIVATE GATEWAY                                                                  #
##############################################################################################################

# 10.1 CREATE VIRTUAL PRIVATE GATEWAY
def vpn_gateway():
    try:
        zone = input("\nAvailabilityZone : ")
        response = client.create_vpn_gateway(
            AvailabilityZone=zone,
            Type='ipsec.1',
            AmazonSideAsn=64513,
        )
        connectionId = response['VpnGateway']['VpnGatewayId']
        print(connectionId)
        name = input("VPG Name : ")
        client.create_tags(
            Resources=[connectionId],
            Tags=[{"Key": "Name", "Value": name}]
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 10.2 Describe VIRTUAL PRIVATE GATEWAY
def describe_vpn_gateway():
    try:
        response = client.describe_vpn_gateways()
        vpn = response['VpnGateways']
        print("\nThe available VPG in your region\n")
        for vpng in vpn:
            fetch = vpng['VpnGatewayId']
            state = vpng['State']
            tags = vpng['Tags'][0]
            vpc = vpng['VpcAttachments'][0]
            print("\t\tVPN Gateway-ID : {} | State : {} | VPN Gateway-NAME : {} | Attached VPC : {}".
                  format(fetch, state, tags['Value'], vpc['VpcId']))
    except ClientError as e:
        print("\n")
        print(e)


# 10.3 ATTACH VIRTUAL PRIVATE GATEWAY
def attach_vpn_gateway():
    try:
        describe_vpc()
        describe_vpn_gateway()
        id = input("\nVPC-ID : ")
        vpg_id = input("VpnGatewayId : ")
        response = client.attach_vpn_gateway(
            VpcId=id,
            VpnGatewayId=vpg_id,
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 10.4 DETACH VIRTUAL PRIVATE GATEWAY
def detach_vpn_gateway():
    try:
        describe_vpc()
        describe_vpn_gateway()
        id = input("\nVpcID : ")
        vpg_id = input("VpnGatewayId : ")
        response = client.detach_vpn_gateway(
            VpcId=id,
            VpnGatewayId=vpg_id,
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 10.5 DELETE VIRTUAL PRIVATE GATEWAY
def delete_vpn_gateway():
    try:
        describe_vpn_gateway()
        vpg_id = input("\nVpnGatewayId : ")
        response = client.delete_vpn_gateway(
            VpnGatewayId=vpg_id,
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 11.SITE-TO-SITE VPN CONNECTION
##############################################################################################################
#       11.1 CREATE SITE-TO-SITE VPN CONNECTION                                                             #
#       11.2 Describe SITE-TO-SITE VPN CONNECTION                                                           #
#       11.3 CREATE STATIC ROUTE IN SITE-TO-SITE VPN CONNECTION                                             #
#       11.4 DELETE STATIC ROUTE IN SITE-TO-SITE VPN CONNECTION                                             #
#       11.5 MODIFY CUSTOMER GATEWAY IN SITE-TO-SITE VPN CONNECTION                                         #
#       11.6 MODIFY VIRTUAL PRIVATE GATEWAY IN SITE-TO-SITE VPN CONNECTION                                  #
#       11.7 DELETE SITE-TO-SITE VPN CONNECTION                                                             #
##############################################################################################################

# 11.1 CREATE SITE-TO-SITE VPN CONNECTION
def vpn_connection():
    try:
        describe_customer_gateway()
        describe_vpn_gateway()
        cid = input("\nCustomerGatewayId : ")
        vid = input("VpnGatewayId : ")
        response = client.create_vpn_connection(
            CustomerGatewayId=cid,
            Type='ipsec.1',
            VpnGatewayId=vid,
            Options={
                'StaticRoutesOnly': True
            }
        )
        connectionId = response['VpnConnection']['VpnConnectionId']
        print(connectionId)
        name = input("VPN Name : ")
        client.create_tags(
            Resources=[connectionId],
            Tags=[{"Key": "Name", "Value": name}]
        )

        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 11.2 Describe SITE-TO-SITE VPN CONNECTION
def describe_vpn_connections():
    try:
        response = client.describe_vpn_connections()
        vpn = response['VpnConnections']
        print("\nThe available Site-to-Site VPN in your region\n")
        for svpn in vpn:
            fetch = svpn['VpnConnectionId']
            tag = svpn['Tags'][0]
            state = svpn['State']
            print("\t\tSite-to-Site VPN ID : {} | Name : {} | State : {}".format(fetch, tag['Value'], state))
    except ClientError as e:
        print("\n")
        print(e)


# 11.3 CREATE STATIC ROUTE IN SITE-TO-SITE VPN CONNECTION
def vpn_connection_static_route():
    try:
        describe_vpn_connections()
        cidr = input("\nDestinationCidrBlock : ")
        id = input("VpnConnectionId : ")
        response = client.create_vpn_connection_route(
            DestinationCidrBlock=cidr,
            VpnConnectionId=id
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 11.4 DELETE STATIC ROUTE IN SITE-TO-SITE VPN CONNECTION
def delete_vpn_connection_static_route():
    try:
        describe_vpn_connections()
        cidr = input("\nDestinationCidrBlock : ")
        id = input("VpnConnectionId : ")
        response = client.delete_vpn_connection_route(
            DestinationCidrBlock=cidr,
            VpnConnectionId=id
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 11.5 MODIFY CUSTOMER GATEWAY IN SITE-TO-SITE VPN CONNECTION
def modify_customer_gateway_vpn_connection():
    try:
        describe_vpn_connections()
        describe_customer_gateway()
        vpn_id = input("\nVpnConnectionId : ")
        cid = input("CustomerGatewayId : ")
        response = client.modify_vpn_connection(
            VpnConnectionId=vpn_id,
            CustomerGatewayId=cid,
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# 11.6 MODIFY VIRTUAL PRIVATE GATEWAY IN SITE-TO-SITE VPN CONNECTION
def modify_virtual_private_gateway_vpn_connection():
    try:
        describe_vpn_connections()
        describe_vpn_gateway()
        vpn_id = input("\nVpnConnectionId : ")
        vpg_id = input("VpnGatewayId : ")
        response = client.modify_vpn_connection(
            VpnConnectionId=vpn_id,
            VpnGatewayId=vpg_id,
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


#  11.7 DELETE SITE-TO-SITE VPN CONNECTION
def delete_vpn_connection():
    try:
        describe_vpn_connections()
        vpn_id = input("\nVpnConnectionId : ")
        response = client.delete_vpn_connection(
            VpnConnectionId=vpn_id,
        )
        print("\nSuccess")
    except ClientError as e:
        print('\n')
        print(e)


# 12.VPC ENDPOINTS
##############################################################################################################
#       12.1 CREATE ENDPOINTS                                                                                #
#       12.2 Describe ENDPOINTS                                                                              #
#       12.3 DELETE ENDPOINTS                                                                                #
##############################################################################################################

# 12.1 CREATE ENDPOINTS
def endpoints():
    try:
        describe_vpc()
        print("\nService Names")
        print("""
        \tcom.amazonaws.eu-west-1.secretsmanager
        \tcom.amazonaws.eu-west-1.ec2
        \tcom.amazonaws.eu-west-1.monitoring
        \tcom.amazonaws.eu-west-1.logs
        \tcom.amazonaws.eu-west-1.events
        """)
        describe_subnet()
        describe_route_table()
        describe_security_group()
        print("""

   **** NOTE FOR ENDPOINT TYPE ****
        
while attaching SUBNETS use --> Interface

while attaching ROUTETABLE use --> Gateway
""")
        type = input("\nVpc Endpoint Type -> {Interface | Gateway} : ")
        vpc_id = input("\nVPC ID : ")
        ser_name = input("\nServiceName : ")
        if type == 'Interface':
            subnets = input("\nEnter the Subnets need to attach (Enter mutliple keywords seperate by ','): ")
            subnets_list = subnets.split(",")
            security_group = input("\nEnter the Security Groups need to attach (Enter mutliple keywords seperate by ','): ")
            security_group_list = security_group.split(",")
        else:
            route_table = input("\nEnter the Route Table need to attach (Enter mutliple keywords seperate by ','): ")
            route_table_list = route_table.split(",")

        response = client.modify_vpc_attribute(
            EnableDnsHostnames={
        'Value': True
        },
            VpcId=vpc_id
        )

        if type == 'Interface':
            response = client.create_vpc_endpoint(
                VpcEndpointType=type,
                VpcId=vpc_id,
                ServiceName=ser_name,
                SubnetIds=subnets_list,
                SecurityGroupIds=security_group_list,
            )
            ConnectionId = response['VpcEndpoint']['VpcEndpointId']
            print(ConnectionId)
            name = input("VPC Endpoint Name : ")
            client.create_tags(
                Resources=[ConnectionId],
                Tags=[{"Key": "Name", "Value": name}]
            )
            print("\nSuccess\n")
        else:
            response = client.create_vpc_endpoint(
                VpcEndpointType=type,
                VpcId=vpc_id,
                ServiceName=ser_name,
                RouteTableIds=route_table_list,
                #SecurityGroupIds=security_group_list,
            )
            ConnectionId = response['VpcEndpoint']['VpcEndpointId']
            print(ConnectionId)
            name = input("VPC Endpoint Name : ")
            client.create_tags(
                Resources=[ConnectionId],
                Tags=[{"Key": "Name", "Value": name}]
            )
            print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)
        

# 12.2 Describe ENDPOINTS
def describe_endpoints():
    try:
        response = client.describe_vpc_endpoints()
        endpoints = response['VpcEndpoints']
        print("\nThe available VPC Endpoints in your region\n")
        for ep in endpoints:
            fetch = ep['VpcEndpointId']
            vpc = ep['VpcId']
            rt = ep['RouteTableIds']
            tags = ep['Tags'][0]
            print("\t\tVPC ENDPOINT-ID : {} | ALLOCATED VPC : {} | ALLOCATED ROUTETABLE : {} | ENDPOINT-NAME : {}".format(fetch, vpc, rt, tags['Value']))
    except ClientError as e:
        print("\n")
        print(e)


# 12.3 DELETE ENDPOINTS
def delete_endpoints():
    try:
        describe_endpoints()
        end_lst = []
        n = int(input("\nNumber of VPC_ENDPOINT need to delete : "))
        for i in range(0, n):
            ele = (input()) 
          
            end_lst.append(ele)
        response = client.delete_vpc_endpoints(
            VpcEndpointIds=end_lst
        )
        print("\nSuccess\n")
    except ClientError as e:
        print("\n")
        print(e)


# TO DESCRIBE SECURITY GROUP
def describe_security_group():
    try:
        response = client.describe_security_groups()
        sec_grp = response['SecurityGroups']
        print("\nThe available Security Group in your region\n")
        for sec in sec_grp:
            fetch = sec['GroupId']
            tags = sec['Tags'][0]
            vpc = sec['VpcId']
            print("\t\tSecurity Group-ID : {} | Name : {} | Attached VPC : {}".format(fetch, tags['Value'], vpc))
    except ClientError as e:
        print("\n")
        print(e)



###################################################################################################################
#       12. INVENTORY                                                                                             #
###################################################################################################################

headers = ['REGION', 'VPC NAME', 'VPC ID', 'VPC CIDR', 'SUBNET Name', 'SUBNET ID', 'SUBNET CIDR', 'SUBNET ZONE', 'ROUTETABLE ID', 'ROUTETABLE NAME',
           'ENDPOINT ID', 'ENDPOINT NAME', 'PEERING ID', 'PEERING NAME', 'REQUESTER VPC-ID', 'ACCEPTER VPC-ID', 'CUSTOMERGATEWAY NAME', 'CUSTOMERGATEWAY ID',
           'VIRTUALPRIVATEGATEWAY NAME', 'VIRTUALPRIVATEGATEWAY ID', 'VPN NAME', 'VPN ID']


def describe_inventory():
    # Output
    html_start = "<!DOCTYPE html><html><head><style>table {font-family: arial, sans-serif;border-collapse: collapse;width: 100%;}td, th {border: 1px solid #dddddd;text-align: left;padding: 8px;}tr:nth-child(even) {background-color: #dddddd;}</style></head><body>"
    html_end = "</body></html>"
    file = open('result.html', 'wt')
    file.write(html_start)

    #excel
    workbook = load_workbook("results.xlsx")
    sheet = workbook.create_sheet(rg)
    font = Font(bold = True)
    vpc_font = Font(bold = True, color = "00009688")

    #cells
    vpc_id_cell = sheet.cell(row=1,column=1)
    vpc_id_cell.value = "VPC ID"
    vpc_id_cell.font = font

    vpc_name_cell = sheet.cell(row=1,column=2)
    vpc_name_cell.value = "VPC Name"
    vpc_name_cell.font = font

    vpc_cidr_cell = sheet.cell(row=1,column=3)
    vpc_cidr_cell.value = "VPC CIDR"
    vpc_cidr_cell.font = font

    subnet_cell = sheet.cell(row=1,column=4)
    subnet_cell.value = "SUBNET ID"
    subnet_cell.font = font

    subnet_name_cell = sheet.cell(row=1,column=5)
    subnet_name_cell.value = "SUBNET Name"
    subnet_name_cell.font = font

    subnet_cidr_cell = sheet.cell(row=1,column=6)
    subnet_cidr_cell.value = "SUBNET CIDR"
    subnet_cidr_cell.font = font

    subnet_zone_cell = sheet.cell(row=1,column=7)
    subnet_zone_cell.value = "SUBNET ZONE"
    subnet_zone_cell.font = font

    route_table_cell = sheet.cell(row=1,column=8)
    route_table_cell.value = "ROUTETABLE ID"
    route_table_cell.font = font

    route_name_cell = sheet.cell(row=1,column=9)
    route_name_cell.value = "ROUTETABLE Name"
    route_name_cell.font = font

    endpoint_cell = sheet.cell(row=1,column=10)
    endpoint_cell.value = "ENDPOINT ID"
    endpoint_cell.font = font

    endpoint_name = sheet.cell(row=1,column=11)
    endpoint_name.value = "ENDPOINT NAME"
    endpoint_name.font = font

    peering_cell = sheet.cell(row=1,column=12)
    peering_cell.value = "PEEING ID"
    peering_cell.font = font

    peering_name_cell = sheet.cell(row=1,column=13)
    peering_name_cell.value = "PEERING Name"
    peering_name_cell.font = font

    requester_cell = sheet.cell(row=1,column=14)
    requester_cell.value = "REQUESTER VPC-ID"
    requester_cell.font = font

    accepter_cell = sheet.cell(row=1,column=15)
    accepter_cell.value = "ACCEPTER VPC-ID"
    accepter_cell.font = font

    customer_gateway_cell = sheet.cell(row=1,column=16)
    customer_gateway_cell.value = "CUSTOMERGATEWAY ID"
    customer_gateway_cell.font = font

    customer_gateway_name_cell = sheet.cell(row=1,column=17)
    customer_gateway_name_cell.value = "CUSTOMERGATEWAY Name"
    customer_gateway_name_cell.font = font

    customer_gateway_ipaddress_cell = sheet.cell(row=1,column=18)
    customer_gateway_ipaddress_cell.value = "CUSTOMERGATEWAY IpAddress"
    customer_gateway_ipaddress_cell.font = font

    vpg_cell = sheet.cell(row=1,column=19)
    vpg_cell.value = "VPG ID"
    vpg_cell.font = font

    vpg_name_cell = sheet.cell(row=1,column=20)
    vpg_name_cell.value = "VPG Name"
    vpg_name_cell.font = font

    vpn_cell = sheet.cell(row=1,column=21)
    vpn_cell.value = "VPN ID"
    vpn_cell.font = font

    vpn_name_cell = sheet.cell(row=1,column=22)
    vpn_name_cell.value= "VPN Name"
    vpn_name_cell.font = font

    
    
    #row_count
    row_count = 0;
    
    # VPC
    response = client.describe_vpcs()
    vpcs = response['Vpcs']
    vpc_count = 0;
    for vpc in vpcs:

        row_count = sheet.max_row;

        # Header
        x = PrettyTable()
        x.field_names = headers

        region = rg
        x.add_row([region, "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])

        vpc_id = vpc['VpcId']
        vpc_name = vpc['Tags'][0]
        vpc_cidr = vpc['CidrBlock']

        x.add_row(["", vpc_name['Value'], vpc_id, vpc_cidr, "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        
        vpc_count = vpc_count + 1;
        # For ID
        id_cell = sheet.cell(row=row_count+vpc_count,column=1)
        id_cell.value = vpc_id
        id_cell.font = vpc_font
        # For Name
        name_cell = sheet.cell(row=row_count+vpc_count,column=2)
        name_cell.value = vpc_name['Value']
        # For CIDR
        cidr_cell = sheet.cell(row=row_count+vpc_count,column=3)
        cidr_cell.value = vpc_cidr
        

        # Subnet
        response = client.describe_subnets(Filters=[{
            "Name": "vpc-id",
            "Values": [vpc_id]
        }])
        subnets = response['Subnets']
        subnet_id_count = 0;
        for subnet in subnets:
            subnet_id = subnet['SubnetId']
            subnet_name = subnet['Tags'][0]
            subnet_cidr = subnet['CidrBlock']
            subnet_zone = subnet['AvailabilityZone']          
            x.add_row(["", "", "", "", subnet_name['Value'], subnet_id, subnet_cidr, subnet_zone, "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
            # For ID
            subnet_id_count = subnet_id_count + 1;
            subnet_cell = sheet.cell(row=row_count+subnet_id_count,column=4)
            subnet_cell.value=subnet_id
            # For Name
            name_cell = sheet.cell(row=row_count+subnet_id_count,column=5)
            name_cell.value=subnet_name['Value']
            # For CIDR
            cidr_cell = sheet.cell(row=row_count+subnet_id_count,column=6)
            cidr_cell.value = subnet_cidr
            # For AvailbitityZone
            zone_cell = sheet.cell(row=row_count+subnet_id_count,column=7)
            zone_cell.value = subnet_zone

        # Route table
        response = client.describe_route_tables(Filters=[{
            "Name": "vpc-id",
            "Values": [vpc_id]
        }])
        route_tables = response['RouteTables']
        route_tables_count = 0;
        for route_table in route_tables:
            route_table_id = route_table['RouteTableId']
            route_table_name = route_table['Tags'][0]
            x.add_row(
                ["", "", "", "", "", "", "", "", route_table_id, route_table_name['Value'], "", "", "", "", "", "", "", "","", "", "", ""])
            # For ID
            route_tables_count = route_tables_count + 1;
            route_table_cell = sheet.cell(row=row_count+route_tables_count,column=8)
            route_table_cell.value = route_table_id
            # For Name
            name_cell = sheet.cell(row=row_count+route_tables_count,column=9)
            name_cell.value=route_table_name['Value']


        # ENDPOINTS
        response = client.describe_vpc_endpoints(Filters=[{
            "Name": "vpc-id",
            "Values": [vpc_id]
        }])
        endpoint = response['VpcEndpoints']
        endpoint_count = 0;
        for ep in endpoint:
            endpoint_id = ep['VpcEndpointId']
            endpoint_name = ep['Tags'][0]
            x.add_row(["", "", "", "", "", "", "", "", "", "", endpoint_id, endpoint_name['Value'], "", "", "", "", "", "", "", "", "", ""])

            # For ID
            endpoint_count = endpoint_count+1;
            endpoint_cell = sheet.cell(row=row_count+endpoint_count,column=10)
            endpoint_cell.value=endpoint_id
            # For Name
            name_cell = sheet.cell(row=row_count+endpoint_count,column=11)
            name_cell.value=endpoint_name['Value']


        # PEERING
        response = client.describe_vpc_peering_connections(Filters=[{
            "Name": "accepter-vpc-info.vpc-id",
            "Values": [vpc_id]
        }])
        #pprint(response)
        #for r in response['VpnConnections']:
        peering = response['VpcPeeringConnections']
        peering_count = 0;
        for vpc_peering in peering:
            vpc_peering_id = vpc_peering['VpcPeeringConnectionId']
            vpc_peering_name = vpc_peering['Tags'][0]
            accepter_vpc_id = vpc_peering['AccepterVpcInfo']['VpcId']
            requester_vpc_id = vpc_peering['RequesterVpcInfo']['VpcId']
            x.add_row(["", "", "", "", "", "", "", "", "", "", "", "",vpc_peering_id, vpc_peering_name['Value'],
                       requester_vpc_id, accepter_vpc_id, "", "", "", "", "", ""])
            # For ID
            peering_count = peering_count + 1;
            peering_cell = sheet.cell(row=row_count+peering_count,column=12)
            peering_cell.value=vpc_peering_id
            # For Name
            name_cell = sheet.cell(row=row_count+peering_count,column=13)
            name_cell.value=vpc_peering_name['Value']
            # For Requester VPC-ID
            requester_cell = sheet.cell(row=row_count+peering_count,column=14)
            requester_cell.value=requester_vpc_id
            # For Accepter VPC-ID
            accepter_cell = sheet.cell(row=row_count+peering_count,column=15)
            accepter_cell.value=accepter_vpc_id

        response_1 = client.describe_vpc_peering_connections(Filters=[{
            "Name": "requester-vpc-info.vpc-id",
            "Values": [vpc_id]
        }])
        #pprint(response_1)
        peering_1 = response_1['VpcPeeringConnections']
        peering_count = 0;
        for vpc_peering_1 in peering_1:
            vpc_peering_id = vpc_peering_1['VpcPeeringConnectionId']
            vpc_peering_name = vpc_peering_1['Tags'][0]
            accepter_vpc_id = vpc_peering_1['AccepterVpcInfo']['VpcId']
            requester_vpc_id = vpc_peering_1['RequesterVpcInfo']['VpcId']

            # For ID
            peering_count = peering_count + 1;
            peering_cell = sheet.cell(row=row_count+peering_count,column=12)
            peering_cell.value=vpc_peering_id
            # For Name
            name_cell = sheet.cell(row=row_count+peering_count,column=13)
            name_cell.value=vpc_peering_name['Value']
            # For Requester VPC-ID
            requester_cell = sheet.cell(row=row_count+peering_count,column=14)
            requester_cell.value=requester_vpc_id
            # For Accepter VPC-ID
            accepter_cell = sheet.cell(row=row_count+peering_count,column=15)
            accepter_cell.value=accepter_vpc_id


        # VPN
        response = client.describe_vpn_gateways()
        virtual_private_gateways = response['VpnGateways']
        for vpg in virtual_private_gateways:
            virtual_private_gateway_id = vpg['VpnGatewayId']
            vpc_attachments = vpg['VpcAttachments']
            
            vpn_tables_count = 0;
            for va in vpc_attachments:
                if va['VpcId'] == vpc_id:
                    # VPN & Customer Gateway
                    response = client.describe_vpn_connections(Filters=[{
                        'Name': 'vpn-gateway-id',
                        'Values': [virtual_private_gateway_id]
                    }])
                    #pprint(response)
                    #print(len(response['VpnConnections']))
                    for r in response['VpnConnections']:
                        vpn_connections=r
                        # vpn_connections = response['VpnConnections'][0]
                        vpn_connection_id = vpn_connections['VpnConnectionId']
                        vpn_connection_name = vpn_connections['Tags'][0]['Value']
                        customer_gateway_id = vpn_connections['CustomerGatewayId']

                        responce = client.describe_vpn_gateways(VpnGatewayIds=[virtual_private_gateway_id])
                        virtual_private_gateway_name = responce['VpnGateways'][0]['Tags'][0]['Value']

                        response = client.describe_customer_gateways(CustomerGatewayIds=[customer_gateway_id])
                        customer_gateway_name = response['CustomerGateways'][0]['Tags'][0]['Value']
                        customer_gateway_ip = response['CustomerGateways'][0]['IpAddress']

                        x.add_row(
                            ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", customer_gateway_id, customer_gateway_name,
                             virtual_private_gateway_id, virtual_private_gateway_name, vpn_connection_id, vpn_connection_name])

                        # For ID
                        vpn_tables_count = vpn_tables_count + 1;
                        vpn_table_cell = sheet.cell(row=row_count+vpn_tables_count,column=16)
                        vpn_table_cell.value = customer_gateway_id
                        # For Name
                        name_cell = sheet.cell(row=row_count+vpn_tables_count,column=17)
                        name_cell.value = customer_gateway_name
                        # For IpAddress
                        vpn_table_cell = sheet.cell(row=row_count+vpn_tables_count,column=18)
                        vpn_table_cell.value = customer_gateway_ip

                        # For ID
                        vpn_table_cell = sheet.cell(row=row_count+vpn_tables_count,column=19)
                        vpn_table_cell.value = virtual_private_gateway_id
                        # For Name
                        name_cell = sheet.cell(row=row_count+vpn_tables_count,column=20)
                        name_cell.value = virtual_private_gateway_name

                        # For ID
                        vpn_table_cell = sheet.cell(row=row_count+vpn_tables_count,column=21)
                        vpn_table_cell.value = vpn_connection_id
                        # For Name
                        name_cell = sheet.cell(row=row_count+vpn_tables_count,column=22)
                        name_cell.value = vpn_connection_name



        vpc_count = 0;
            

        html = x.get_html_string()
        file.write(html)

    file.write(html_end)
    file.close()
    #webbrowser.open('result.html')

    workbook.save("results.xlsx")



def describe_route_inventory():
    #excel
    workbook = load_workbook("routes.xlsx")
    sheet = workbook.create_sheet(rg)
    font = Font(bold = True)
    vpc_font = Font(bold = True, color = "00009688")

    #cells
    vpc_id_cell = sheet.cell(row=1,column=1)
    vpc_id_cell.value = "VPC ID"
    vpc_id_cell.font = font

    vpc_name_cell = sheet.cell(row=1,column=2)
    vpc_name_cell.value = "VPC Name"
    vpc_name_cell.font = font

    vpc_cidr_cell = sheet.cell(row=1,column=3)
    vpc_cidr_cell.value = "VPC CIDR"
    vpc_cidr_cell.font = font

    route_table_cell = sheet.cell(row=1,column=4)
    route_table_cell.value = "ROUTETABLE ID"
    route_table_cell.font = font

    route_name_cell = sheet.cell(row=1,column=5)
    route_name_cell.value = "ROUTETABLE Name"
    route_name_cell.font = font

    local_cidr_cell = sheet.cell(row=1,column=6)
    local_cidr_cell.value = "CIDR"
    local_cidr_cell.font = font

    local_id_cell = sheet.cell(row=1,column=7)
    local_id_cell.value = "TARGET"
    local_id_cell.font = font


    

    #row_count
    row_count = 0;
    
    # VPC
    response = client.describe_vpcs()
    vpcs = response['Vpcs']
    vpc_count = 0;
    for vpc in vpcs:

        row_count = sheet.max_row;

        # Header
        x = PrettyTable()
        #x.field_names = headers

        region = rg
        x.add_row([region, "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])

        vpc_id = vpc['VpcId']
        vpc_name = vpc['Tags'][0]
        vpc_cidr = vpc['CidrBlock']

        x.add_row(["", vpc_name['Value'], vpc_id, vpc_cidr, "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        
        vpc_count = vpc_count + 1;
        # For ID
        id_cell = sheet.cell(row=row_count+vpc_count,column=1)
        id_cell.value = vpc_id
        id_cell.font = vpc_font
        # For Name
        name_cell = sheet.cell(row=row_count+vpc_count,column=2)
        name_cell.value = vpc_name['Value']
        # For CIDR
        cidr_cell = sheet.cell(row=row_count+vpc_count,column=3)
        cidr_cell.value = vpc_cidr

        # Route table
        response = client.describe_route_tables(Filters=[{
            "Name": "vpc-id",
            "Values": [vpc_id]
        }])
        #pprint(response)
        route_tables = response['RouteTables']
        route_tables_count = 0;
        for association in route_tables:
            rt_id = association['RouteTableId']
            rt_name = association['Tags'][0]

            for route in association['Routes']:

                # For ID
                route_table_id = sheet.cell(row=row_count+1,column=4)
                route_table_id.value = rt_id
                # For Name
                route_table_name = sheet.cell(row=row_count+1,column=5)
                route_table_name.value = rt_name['Value']

                if 'GatewayId' in route:

                    route_tables_count = route_tables_count +1;

                    # For LOCAL Route-CIDR
                    cidr = sheet.cell(row=row_count+route_tables_count,column=6)
                    cidr.value = route['DestinationCidrBlock']
                    # For LOCAL Route-CIDR
                    target = sheet.cell(row=row_count+route_tables_count,column=7)
                    target.value = route['GatewayId']


                if 'EgressOnlyInternetGatewayId' in route:
                    route_tables_count = route_tables_count +1;
                    egress_igw_cidr = route['DestinationCidrBlock']
                    egress_igw_id = route['EgressOnlyInternetGatewayId']

                    # For Peering Route-CIDR
                    egress_igw_route_cidr = sheet.cell(row=row_count+route_tables_count,column=6)
                    egress_igw_route_cidr.value = egress_igw_cidr
                    # For Peering Route-ID
                    egress_igw_route_id = sheet.cell(row=row_count+route_tables_count,column=7)
                    egress_igw_route_id.value = egress_igw_id 

                
                if 'InstanceId' in route:
                    route_tables_count = route_tables_count +1;
                    instance_cidr = route['DestinationCidrBlock']
                    instance_id = route['InstanceId']

                    # For INSTANCE Route-CIDR
                    instance_route_cidr = sheet.cell(row=row_count+route_tables_count,column=6)
                    instance_route_cidr.value = instance_cidr
                    # For INSTANCE Route-CIDR
                    instance_route_id = sheet.cell(row=row_count+route_tables_count,column=7)
                    instance_route_id.value = instance_id

                if 'NatGatewayId' in route:
                    route_tables_count = route_tables_count +1;
                    natgateway_cidr = route['DestinationCidrBlock']
                    natgateway_id = route['NatGatewayId']
                    
                    # For INSTANCE Route-CIDR
                    natgateway_route_cidr = sheet.cell(row=row_count+route_tables_count,column=6)
                    natgateway_route_cidr.value = natgateway_cidr
                    # For INSTANCE Route-CIDR
                    natgateway_route_id = sheet.cell(row=row_count+route_tables_count,column=7)
                    natgateway_route_id.value = natgateway_id
                    
                if 'VpcPeeringConnectionId' in route:
                    route_tables_count = route_tables_count +1;
                    peering_cidr = route['DestinationCidrBlock']
                    peering_id = route['VpcPeeringConnectionId']

                    # For Peering Route-CIDR
                    peering_route_cidr = sheet.cell(row=row_count+route_tables_count,column=6)
                    peering_route_cidr.value = peering_cidr
                    # For Peering Route-ID
                    peering_route_id = sheet.cell(row=row_count+route_tables_count,column=7)
                    peering_route_id.value = peering_id

                if 'TransitGatewayId' in route:
                    route_tables_count = route_tables_count +1;
                    transistgateway_cidr = route['DestinationCidrBlock']
                    transistgateway_id = route['TransitGatewayId']

                    # For Peering Route-CIDR
                    transistgateway_route_cidr = sheet.cell(row=row_count+route_tables_count,column=6)
                    transistgateway_route_cidr.value = transistgateway_cidr
                    # For Peering Route-ID
                    transistgateway_route_id = sheet.cell(row=row_count+route_tables_count,column=7)
                    transistgateway_route_id.value = transistgateway_id

                if 'LocalGatewayId' in route:
                    route_tables_count = route_tables_count +1;
                    localgateway_cidr = route['DestinationCidrBlock']
                    localgateway_id = route['LocalGatewayId']

                    # For Peering Route-CIDR
                    localgateway_route_cidr = sheet.cell(row=row_count+route_tables_count,column=6)
                    localgateway_route_cidr.value = localgateway_cidr
                    # For Peering Route-ID
                    localgateway_route_id = sheet.cell(row=row_count+route_tables_count,column=7)
                    localgateway_route_id.value = localgateway_id

                if 'NetworkInterfaceId' in route:
                    route_tables_count = route_tables_count +1;
                    networkinterface_cidr = route['DestinationCidrBlock']
                    networkinterface_id = route['NetworkInterfaceId']

                    # For Peering Route-CIDR
                    networkinterface_route_cidr = sheet.cell(row=row_count+route_tables_count,column=6)
                    networkinterface_route_cidr.value = networkinterface_cidr
                    # For Peering Route-ID
                    networkinterface_route_id = sheet.cell(row=row_count+route_tables_count,column=7)
                    networkinterface_route_id.value = networkinterface_id


        vpc_count = 0;


    workbook.save("routes.xlsx")


##############################################################################################################
#       13. DECLARING IN ARGPARSER                                                                           #
##############################################################################################################


parser = argparse.ArgumentParser(description='AWS',
                                 usage="""

CONNECTION : 
1.TO ESTABLISH CONNECTION -> connection 
NOTE : WITHOUT USING CONNECTION YOU CAN'T DO ANYTHING

INVENTORY : 
1.TO SEE INVENTORY -> inventory

VPC : 
1.TO CREATE VPC -> vpc 
2.TO DESCRIBE VPC -> created_vpc 
3.TO ADD EXTRA CIDR -> add_cidr 
4.TO DELETE VPC -> delete_vpc

SUBNET : 
1.TO CREATE SUBNET -> subnet 
2.TO DESCRIBE SUBNET -> created_subnet 
3.TO DELETE SUBNET-> delete_subnet

ROUTE TABLE : 
1.TO CREATE ROUTE TABLE -> route_table 
2.TO UPDATE ROUTE TABLE -> update_route_table 
3.TO DESCRIBE ROUTE TABLE -> created_route_table 
4.TO DESCRIBE ROUTE -> created_route 
5.TO DELETE ROUTE -> delete_route 
6.TO DELETE ROUTE TABLE -> delete_route_table
7.CREATE ROUTE TO IGW -> route_to_igw 
8.CREATE ROUTE TO VPC PEERING -> route_to_vpc_peering 
9.CREATE ROUTE TO VPG -> route_to_vpg 
10.TO ATTACH SUBNET IN ROUTE TABLE ->attach_subnet_to_route_table

IGW : 
1.TO CREATE IGW ->igw 
2.TO DESCRIBE IGW -> created_igw 
3.ATTACH IGW WITH VPC -> attach_igw_to_vpc 
4.DETACH IGW WITH VPC -> detach_igw_to_vpc 
5.TO DELETE IGW -> delete_igw

PEERING : 
1.TO CREATE PEERING -> peering
2.TO CREATE PEERING IN ANOTHER ACCOUNT -> peering_to_another_account
3.TO DESCRIBE PEERING -> created_peering 
4.ACCEPT PEERING -> accept_peering 
5.REJECT PEERING -> reject_peering 
6.TO DELETE PEERING -> delete_peering

CUSTOMER GATEWAY : 
1.TO CREATE CUSTOMER GATEWAY -> customer_gateway 
2.TO DESCRIBE CUSTOMER GATEWAY -> created_customer_gateway 
3.TO DELETE CUSTOMER GATEWAY -> delete_customer_gateway

VIRTUALPRIVATE GATEWAY : 
1.TO CREATE VPG -> vpg 
2.TO DESCRIBE VPG ->created_vpg 
3.TO ATTACH VPG WITH VPC -> attach_vpg 
4.TO DETACH VPG WITH VPC -> detach_vpg 
5.TO DELETE VPG -> delete_vpg

VPN : 
1.TO CREATE VPN -> vpn 
2.TO DESCRIBE VPN -> created_vpn 
3.TO CREATE STATIC ROUTE IN VPN -> create_static_route_in_vpn 
4.TO DELETE STATIC ROUTE TABLE -> delete_static_route_in_vpn 
5.TO MODIFY CUSTOMER GATEWAY IN VPN -> modify_customer_gateway_in_vpn 
6.TO MODIFY VPG IN VPN ->modify_vpg_in_vpn 
7.TO DELETE VPN -> delete_vpn

ENDPOINTS :
1.TO CREATE ENDPOINTS -> endpoint
2.TO DESCRIBE ENDPOINT -> created_endpoint
3.TO DELETE ENDPOINT -> delete_endpoint

INVENTORY
1.FOR TOTAL INVENTORY -> inventory
2.FOR ROUTE TABLE INVENTORY -> inventory_routetable

""")

parser.add_argument('action', choices=['connection'])
parser.add_argument('function', choices=[
    'vpc',
    'created_vpc',
    'add_cidr',
    'delete_vpc',
    'subnet',
    'created_subnet',
    'delete_subnet',
    'route_table',
    'route_to_igw',
    'created_route_table',
    'created_route',
    'delete_route',
    'delete_route_table',
    'update_route_table',
    'route_to_vpc_peering',
    'route_to_vpg',
    'attach_subnet_to_route_table',
    'detach_subnet_to_route_table',
    'igw',
    'created_igw',
    'attach_igw_to_vpc',
    'detach_igw_to_vpc',
    'delete_igw',
    'peering',
    'peering_to_another_account',
    'created_peering',
    'accept_peering',
    'reject_peering',
    'delete_peering',
    'customer_gateway',
    'created_customer_gateway',
    'delete_customer_gateway',
    'vpg',
    'created_vpg',
    'attach_vpg',
    'detach_vpg',
    'delete_vpg',
    'vpn',
    'created_vpn',
    'create_static_route_in_vpn',
    'delete_static_route_in_vpn',
    'modify_customer_gateway_in_vpn',
    'modify_vpg_in_vpn',
    'delete_vpn',
    'endpoint',
    'created_endpoint',
    'delete_endpoint',
    'inventory',
    'inventory_routetable'
    ])

args = parser.parse_args()

if args.action == "connection":
    region = input("\nEnter the region : ")

    rg = region
    
    client = connection(region)
    clients = connection1(region)
    

if args.function == "vpc":
    vpc()

if args.function == "created_vpc":
    describe_vpc()

if args.function == "add_cidr":
    secondary_cidr()
    
if args.function == "delete_vpc":
    delete_vpc()

if args.function == "subnet":
    subnet()

elif args.function == "created_subnet":
    describe_subnet()

if args.function == "delete_subnet":
    delete_subnet()

if args.function == "route_table":
    route_table()

if args.function == "route_to_igw":
    route_igw()

if args.function == "created_route_table":
    describe_route_table()

if args.function == "created_route":
    describe_route()

if args.function == "delete_route":
    delete_route()

if args.function == "delete_route_table":
    delete_route_table()

if args.function == "update_route_table":
    update_route_table()

if args.function == "route_to_vpc_peering":
    route_vpc_peering()

if args.function == "route_to_vpg":
    route_vpn_gateway()

if args.function == "attach_subnet_to_route_table":
    attach_subnet_route_table()

if args.function == "detach_subnet_to_route_table":
    detach_subnet_route_table()

if args.function == "igw":
    internet_gateway()

if args.function == "created_igw":
    describe_igw()

if args.function == "attach_igw_to_vpc":
    attach_igw()

if args.function == "detach_igw_to_vpc":
    detach_igw()

if args.function == "delete_igw":
    delete_igw()

if args.function == "peering":
    vpc_peering()

if args.function == "peering_to_another_account":
    vpc_peering_another_region()

if args.function == "created_peering":
    describe_vpc_peering()

if args.function == "accept_peering":
    accept_peering()

if args.function == "reject_peering":
    reject_peering()

if args.function == "delete_peering":
    delete_vpc_peering()

if args.function == "customer_gateway":
    customer_gateway()

if args.function == "created_customer_gateway":
    describe_vpn_gateway()

if args.function == "delete_customer_gateway":
    delete_customer_gateway()


if args.function == "vpg":
    vpn_gateway()

if args.function == "created_vpg":
    describe_vpn_gateway()

if args.function == "attach_vpg":
    attach_vpn_gateway()

if args.function == "detach_vpg":
    detach_vpn_gateway()

if args.function == "delete_vpg":
    delete_vpn_gateway()

if args.function == "vpn":
    vpn_connection()

if args.function == "created_vpn":
    describe_vpn_connections()

if args.function == "create_static_route_in_vpn":
    vpn_connection_static_route()

if args.function == "delete_static_route_in_vpn":
    delete_vpn_connection_static_route()

if args.function == "modify_customer_gateway_in_vpn":
    modify_customer_gateway_vpn_connection()

if args.function == "modify_vpg_in_vpn":
    modify_virtual_private_gateway_vpn_connection()

if args.function == "delete_vpn":
    delete_vpn_connection()

if args.function == 'endpoint':
    endpoints()

if args.function == 'created_endpoint':
    describe_endpoints()

if args.function == 'delete_endpoint':
    delete_endpoints()

if args.function == "inventory":
    describe_inventory()

if args.function == "inventory_routetable":
    describe_route_inventory()
    

print('\nLeaving the script.. Have a great day ahead 😊')



